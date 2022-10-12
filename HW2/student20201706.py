#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

row_id = 1
for row in ws:
	if row_id != 1:
		sum_all = ws.cell(row = row_id, column = 3).value * 0.3
		sum_all += ws.cell(row = row_id, column = 4).value * 0.35
		sum_all += ws.cell(row = row_id, column = 5).value * 0.34
		sum_all += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_all
	row_id += 1
	
stu_num = row_id - 2
#print(stu_num)

gradeA_num = int(stu_num * 0.3)
gradeB_num = int(stu_num * 0.7) - gradeA_num
gradeC_num = stu_num - (gradeA_num + gradeB_num)
gradeAp_num = int(gradeA_num * 0.5)
gradeBp_num = int(gradeB_num * 0.5)
gradeCp_num = int(gradeC_num * 0.5)
gradeA0_num = gradeA_num - gradeAp_num
gradeB0_num = gradeB_num - gradeBp_num
gradeC0_num = gradeC_num - gradeCp_num
gradeF_num = stu_num - (gradeA_num + gradeB_num + gradeC_num)
#print(gradeF_num)

jumsu = []
row_id = 1
for row in ws:
	if row_id != 1:
		jumsu.append(ws.cell(row = row_id, column = 7).value)
	row_id += 1

sorted_jumsu = sorted(jumsu)
sorted_jumsu.reverse()

count = 0
while(gradeAp_num > 0):
	for i in range(count, stu_num):
		for number in jumsu:
			if sorted_jumsu[i] == number:
				row_id = jumsu.index(number) + 2
				ws.cell(row = row_id, column = 8).value = "A+"
				gradeAp_num -= 1
				count += 1
				if gradeAp_num == 0:
					break
		if gradeAp_num == 0:
			break

while(gradeA0_num > 0):
	for i in range(count, stu_num):
		for number in jumsu:
			if sorted_jumsu[i] == number:
				row_id = jumsu.index(number) + 2
				ws.cell(row = row_id, column = 8).value = "A0"
				gradeA0_num -= 1
				count += 1
				if gradeA0_num == 0:
					break
		if gradeA0_num == 0:
			break
	
while(gradeBp_num > 0):
	for i in range(count, stu_num):
		for number in jumsu:
			if sorted_jumsu[i] == number:
				row_id = jumsu.index(number) + 2
				ws.cell(row = row_id, column = 8).value = "B+"
				gradeBp_num -= 1
				count += 1
				if gradeBp_num == 0:
					break
		if gradeBp_num == 0:
			break

while(gradeB0_num > 0):
	for i in range(count, stu_num):
		for number in jumsu:
			if sorted_jumsu[i] == number:
				row_id = jumsu.index(number) + 2
				ws.cell(row = row_id, column = 8).value = "B0"
				gradeB0_num -= 1
				count += 1
				if gradeB0_num == 0:
					break
		if gradeB0_num == 0:
			break

while(gradeCp_num > 0):
	for i in range(count, stu_num):
		for number in jumsu:
			if sorted_jumsu[i] == number:
				row_id = jumsu.index(number) + 2
				ws.cell(row = row_id, column = 8).value = "C+"
				gradeCp_num -= 1
				count += 1
				if gradeCp_num == 0:
					break
		if gradeCp_num == 0:
			break	

while(gradeC0_num > 0):
	for i in range(count, stu_num):
		for number in jumsu:
			if sorted_jumsu[i] == number:
				row_id = jumsu.index(number) + 2
				ws.cell(row = row_id, column = 8).value = "C0"
				gradeC0_num -= 1
				count += 1
				if gradeC0_num == 0:
					break
		if gradeC0_num == 0:
			break	
			
while(gradeF_num > 0):
	for i in range(count, stu_num):
		for number in jumsu:
			if sorted_jumsu[i] == number:
				row_id = jumsu.index(number) + 2
				ws.cell(row = row_id, column = 8).value = "F"
				gradeF_num -= 1
				count += 1
				if gradeF_num == 0:
					break
		if gradeF_num == 0:
			break	
#print(count)
#print(stu_num)
#print(gradeC0_num)
			
wb.save("student.xlsx")	
